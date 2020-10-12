import openpyxl

vk = openpyxl.load_workbook("XLRD.xlsx")

print(vk.sheetnames)

print("Active sheet is:" + vk.active.title)

# create object on any sheet tou wanto to work with

sh = vk['Sheet1']
print(sh.title)
print(sh['A3'].value)

c1 = sh.cell(2,2)
print(c1.value)

#Find rows havving data
rows = sh.max_row
columns = sh.max_column

print("Total rows " + str(rows))
print("Total columns " + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
        c=sh.cell(i,j)
        print(c.value)

# second aproach to fetch data from excel sheet
for r in sh['A1':'C3']:
    for c in r:
        print(c.value)